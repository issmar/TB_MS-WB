import time
import pandas as pd
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

CHROMEDRIVER_PATH = "chromedriver.exe"
PROFILE_PATH = "user-data-dir=C:/Users/ALPHA_PC/AppData/Local/Google/Chrome/User Data/WhatsappProfile"
WA_URL = "https://web.whatsapp.com/"
EXCEL_PATH = "./GENERADOS/contactos_filtrados.xlsx"

def obtener_no_leidos_y_actualizar():
    # Cargar archivo sin perder formato
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # Leer encabezados y encontrar las columnas necesarias
    headers = [cell.value.strip().lower() for cell in ws[1]]
    col_tel = headers.index("telefono.") + 1
    col_estado_wa = headers.index("estado respuesta wa") + 1

    # Crear diccionario {tel_normalizado: fila}
    telefono_dict = {}
    for row in range(2, ws.max_row + 1):
        valor = ws.cell(row=row, column=col_tel).value
        if valor:
            tel = str(valor).strip()
            tel = ''.join(filter(str.isdigit, tel))
            if tel.startswith("52"):
                tel = tel[2:]
            if tel.startswith("1"):
                tel = tel[1:]
            telefono_dict[tel] = row

    # Iniciar navegador
    options = Options()
    options.add_argument(PROFILE_PATH)
    driver = webdriver.Chrome(service=Service(CHROMEDRIVER_PATH), options=options)
    driver.get(WA_URL)

    try:
        WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "pane-side")))
        print("✅ WhatsApp cargado. Esperando a que estés en la sección 'No leídos'...")

        while True:
            btn = driver.find_element(By.ID, "unread-filter")
            if btn.get_attribute("aria-selected") == "true":
                print("📍 Ya estás en la sección 'No leídos'")
                break
            time.sleep(1)

        time.sleep(2)

        chats = driver.find_elements(By.XPATH, '//div[@role="listitem"]')
        print(f"\n📬 Chats encontrados: {len(chats)}")

        for chat in chats:
            try:
                nombre = chat.find_element(By.XPATH, './/span[@title]').get_attribute("title")
                telefono = ''.join(filter(str.isdigit, nombre))

                if telefono.startswith("52"):
                    telefono = telefono[2:]
                if telefono.startswith("1"):
                    telefono = telefono[1:]

                try:
                    no_leidos_span = chat.find_element(By.XPATH, './/span[contains(@aria-label,"no leídos")]')
                    cantidad = no_leidos_span.text.strip()
                except:
                    cantidad = "?"

                print(f"📨 {nombre} → {cantidad} mensajes no leídos")

                if telefono in telefono_dict:
                    fila = telefono_dict[telefono]
                    celda_estado = ws.cell(row=fila, column=col_estado_wa)
                    if str(celda_estado.value).strip().lower() == "pendiente":
                        celda_estado.value = "Respondido"
                        print(f"✅ Estado actualizado en Excel para {nombre}")
            except Exception as e:
                print(f"⚠️ No se pudo procesar un chat: {e}")
                continue

        wb.save(EXCEL_PATH)
        print("\n📁 Excel actualizado con respuestas detectadas.")
        
    except Exception as e:
        print(f"❌ Error: {e}")
    finally:
        input("\nPresiona Enter para cerrar...")
        driver.quit()

if __name__ == "__main__":
    obtener_no_leidos_y_actualizar()
