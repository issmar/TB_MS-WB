import os
import socket
import pandas as pd
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import time
import re

# ==== 1. Diccionario de perfiles ====
perfiles = {
    "Perfil1": {"correo": "ventas.niflorlogistics@gmail.com", "app_password": "letnhuuipprgewsu"},
    "Perfil2": {"correo": "ventas.niflor2@gmail.com", "app_password": "hdjklwmqmvbfqjdo"},
    "Perfil3": {"correo": "ventas.niflor3@gmail.com", "app_password": "yikcbbdvtaldlfxu"},
}

# ==== Validar dominio sintácticamente antes de resolver DNS ====
def dominio_valido(dominio):
    if not dominio or len(dominio) > 253 or ".." in dominio:
        return False
    if "." not in dominio:
        return False
    return re.fullmatch(r"[a-z0-9.-]+", dominio) is not None

# ==== 2. Paso 1: Validar dominios ====
def validar_dominios_y_actualizar_excel(ruta_excel):
    if not os.path.exists(ruta_excel):
        print(f"❌ Archivo no encontrado: {ruta_excel}")
        return

    wb = load_workbook(ruta_excel)
    ws = wb.active

    encabezados = [cell.value for cell in ws[1]]
    col_email = encabezados.index("E Mail") + 1
    col_estado_dominio = encabezados.index("Estado Dominio") + 1
    col_estado_envio = encabezados.index("Estado Envío Correo") + 1
    col_estado_respuesta = encabezados.index("Estado Respuesta Correo") + 1

    print("\n🔍 Validando dominios con estado 'Pendiente'...\n")
    for fila in range(2, ws.max_row + 1):
        estado = str(ws.cell(row=fila, column=col_estado_dominio).value).strip().lower()
        email = str(ws.cell(row=fila, column=col_email).value).strip() if ws.cell(row=fila, column=col_email).value else ""

        if estado != "pendiente":
            continue

        if not email or "@" not in email:
            ws.cell(row=fila, column=col_estado_dominio).value = "No Existe"
            if str(ws.cell(row=fila, column=col_estado_envio).value).strip().lower() == "pendiente":
                ws.cell(row=fila, column=col_estado_envio).value = "No Aplica"
            if str(ws.cell(row=fila, column=col_estado_respuesta).value).strip().lower() == "pendiente":
                ws.cell(row=fila, column=col_estado_respuesta).value = "No Aplica"
            print(f"❌ Fila {fila}: Email inválido o vacío → No Existe / No Aplica")
            continue

        dominio = email.split("@")[-1].strip().lower()

        if not dominio_valido(dominio):
            ws.cell(row=fila, column=col_estado_dominio).value = "No Existe"
            if str(ws.cell(row=fila, column=col_estado_envio).value).strip().lower() == "pendiente":
                ws.cell(row=fila, column=col_estado_envio).value = "No Aplica"
            if str(ws.cell(row=fila, column=col_estado_respuesta).value).strip().lower() == "pendiente":
                ws.cell(row=fila, column=col_estado_respuesta).value = "No Aplica"
            print(f"❌ Fila {fila}: Dominio malformado → '{dominio}' → No Aplica")
            continue

        try:
            socket.gethostbyname(dominio)
            ws.cell(row=fila, column=col_estado_dominio).value = "Existe"
            print(f"✅ Fila {fila}: {email} → Dominio válido")
        except (socket.gaierror, UnicodeError) as e:
            ws.cell(row=fila, column=col_estado_dominio).value = "No Existe"
            if str(ws.cell(row=fila, column=col_estado_envio).value).strip().lower() == "pendiente":
                ws.cell(row=fila, column=col_estado_envio).value = "No Aplica"
            if str(ws.cell(row=fila, column=col_estado_respuesta).value).strip().lower() == "pendiente":
                ws.cell(row=fila, column=col_estado_respuesta).value = "No Aplica"
            print(f"❌ Fila {fila}: {email} → Dominio NO válido ({e}) → No Aplica")

    wb.save(ruta_excel)
    print("\n✅ Validación de dominios completada y guardada en Excel.\n")

# ==== 3. Seleccionar perfil ====
def seleccionar_perfil():
    print("\n📧 Selecciona un perfil para enviar correos:")
    for i, nombre in enumerate(perfiles.keys(), 1):
        print(f"{i}. {nombre} - {perfiles[nombre]['correo']}")
    try:
        seleccion = int(input("Número de perfil: "))
        clave = list(perfiles.keys())[seleccion - 1]
        perfil = perfiles[clave]
        if "pendiente" in perfil["correo"]:
            print("⚠️ Perfil sin credenciales válidas.")
            return None
        return perfil
    except Exception:
        print("❌ Selección inválida.")
        return None

# ==== 4. Paso 2: Enviar correos y actualizar estado ====
def enviar_correos_y_actualizar_excel(ruta_excel, remitente, app_password):
    if not os.path.exists(ruta_excel):
        print("❌ Archivo no encontrado.")
        return

    pdf_filename = "NIFLOR.pdf"
    pdf_path = os.path.abspath(pdf_filename)
    if not os.path.isfile(pdf_path):
        print(f"❌ PDF no encontrado: {pdf_filename}")
        return

    wb = load_workbook(ruta_excel)
    ws = wb.active

    encabezados = [cell.value for cell in ws[1]]
    col_email = encabezados.index("E Mail") + 1
    col_estado_dominio = encabezados.index("Estado Dominio") + 1
    col_estado_envio = encabezados.index("Estado Envío Correo") + 1
    col_correo_envio = encabezados.index("Correo") + 1

    cuerpo_html = """
    <b>🚛 BASE EN ALTAMIRA, TAMPS.</b><br>
    <b>4 TRACTOCAMIONES CON SERVICIO FULL, DISPONIBLES</b><br><br>
    ¿Tienes cargas que no se están moviendo por falta de transporte?<br><br>
    Formada por personas trabajadoras y de confianza, con rutas activas a <b>Nuevo León, San Luis Potosí, Jalisco, Coahuila y Tamaulipas</b>.<br><br>
    Somos una empresa nueva y al tener una flotilla pequeña, podemos darte <b>atención personalizada, disponibilidad inmediata y seguimiento puntual</b>.<br><br>
    ¿Te molesta si cotizamos alguno de tus requerimientos?<br><br>
    Saludos,<br>
    <b>Enrique Delgado</b><br>
    <b>833 236 66 62</b>
    """

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(remitente, app_password)
    except smtplib.SMTPAuthenticationError:
        print("❌ Error de autenticación con SMTP.")
        return
    except Exception as e:
        print(f"❌ Error al conectar al servidor SMTP: {e}")
        return

    print(f"\n📤 Enviando correos desde: {remitente}\n")

    enviados = 0
    for fila in range(2, ws.max_row + 1):
        email = str(ws.cell(row=fila, column=col_email).value).strip()
        estado_dom = str(ws.cell(row=fila, column=col_estado_dominio).value).strip().lower()
        estado_env = str(ws.cell(row=fila, column=col_estado_envio).value).strip().lower()

        if estado_dom != "existe" or estado_env != "pendiente":
            continue

        try:
            msg = MIMEMultipart()
            msg['From'] = remitente
            msg['To'] = email
            msg['Subject'] = "NIFLOR: Servicio de confianza / 4 tractos servicio full"

            msg.attach(MIMEText(cuerpo_html, "html"))

            with open(pdf_path, "rb") as archivo_pdf:
                from email.mime.application import MIMEApplication
                parte_pdf = MIMEApplication(archivo_pdf.read(), _subtype="pdf")
                parte_pdf.add_header("Content-Disposition", "attachment", filename=pdf_filename)
                msg.attach(parte_pdf)

            server.sendmail(remitente, email, msg.as_string())

            ws.cell(row=fila, column=col_estado_envio).value = "Enviado"
            ws.cell(row=fila, column=col_correo_envio).value = remitente
            enviados += 1
            print(f"✅ Enviado a: {email}")
            time.sleep(1)
        except Exception as e:
            print(f"❌ Error al enviar a {email}: {e}")

    wb.save(ruta_excel)
    server.quit()
    print(f"\n✅ Envío completado. Correos enviados: {enviados}\n")

# ==== 5. Ejecutar flujo ====
def main():
    ruta = "./GENERADOS/contactos_filtrados.xlsx"
    validar_dominios_y_actualizar_excel(ruta)
    perfil = None
    while not perfil:
        perfil = seleccionar_perfil()
    enviar_correos_y_actualizar_excel(ruta, perfil["correo"], perfil["app_password"])

if __name__ == "__main__":
    main()
