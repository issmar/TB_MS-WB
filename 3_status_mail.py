import os
import imaplib
import email
import re
from openpyxl import load_workbook

# ==== 1. Diccionario de perfiles ====
perfiles = {
    "Perfil1": {"correo": "ventas.niflorlogistics@gmail.com", "app_password": "letnhuuipprgewsu"},
    "Perfil2": {"correo": "ventas.niflor2@gmail.com", "app_password": "hdjklwmqmvbfqjdo"},
    "Perfil3": {"correo": "ventas.niflor3@gmail.com", "app_password": "yikcbbdvtaldlfxu"},
}

# ==== 2. Verificar por remitente ====
def verificar_estado_por_remitente(ruta_excel, remitente, app_password):
    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(remitente, app_password)
    except Exception as e:
        print(f"\n❌ Error al conectar IMAP con {remitente}: {e}")
        return

    wb = load_workbook(ruta_excel)
    ws = wb.active
    encabezados = [cell.value for cell in ws[1]]

    col_email = encabezados.index("E Mail") + 1
    col_envio = encabezados.index("Estado Envío Correo") + 1
    col_respuesta = encabezados.index("Estado Respuesta Correo") + 1
    col_correo_remitente = encabezados.index("Correo") + 1

    # Obtener filas asociadas a este remitente
    filas_a_verificar = []
    for fila in range(2, ws.max_row + 1):
        remitente_fila = str(ws.cell(row=fila, column=col_correo_remitente).value).strip().lower()
        if remitente_fila == remitente.lower():
            email_dest = str(ws.cell(row=fila, column=col_email).value).strip().lower()
            estado_envio = str(ws.cell(row=fila, column=col_envio).value).strip().lower()
            estado_respuesta = str(ws.cell(row=fila, column=col_respuesta).value).strip().lower()
            if estado_envio in ["enviado", "pendiente"] or estado_respuesta == "pendiente":
                filas_a_verificar.append((fila, email_dest))

    if not filas_a_verificar:
        print(f"ℹ️ No hay correos pendientes para {remitente}")
        return

    print(f"\n📥 Conectado como: {remitente}")
    print(f"🔍 Filas a verificar: {len(filas_a_verificar)}")

    # Buscar rebotes
    carpetas_rebote = ["INBOX", "[Gmail]/Spam", "[Gmail]/Trash"]
    rebotados = set()

    for carpeta in carpetas_rebote:
        try:
            mail.select(carpeta, readonly=True)
            typ, data = mail.search(None, '(FROM "mailer-daemon")')
            for num in data[0].split():
                typ, msg_data = mail.fetch(num, '(RFC822)')
                raw = msg_data[0][1]
                msg = email.message_from_bytes(raw)

                body = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        if part.get_content_type() == "text/plain":
                            body += part.get_payload(decode=True).decode(errors="ignore")
                else:
                    body = msg.get_payload(decode=True).decode(errors="ignore")

                encontrados = re.findall(r'[\w\.-]+@[\w\.-]+', body.lower())
                rebotados.update(encontrados)
        except Exception as e:
            print(f"⚠️ No se pudo revisar carpeta {carpeta}: {e}")

    # Buscar respuestas
    respondidos = set()
    mail.select("INBOX")
    for _, correo_dest in filas_a_verificar:
        search = f'(FROM "{correo_dest}")'
        try:
            typ, data = mail.search(None, search)
            if data[0]:
                respondidos.add(correo_dest)
        except Exception:
            continue

    # Actualizar archivo
    for fila, correo_dest in filas_a_verificar:
        if correo_dest in rebotados:
            ws.cell(row=fila, column=col_envio).value = "Rebotado"
        if correo_dest in respondidos:
            ws.cell(row=fila, column=col_respuesta).value = "Respondido"

    wb.save(ruta_excel)
    print(f"✅ Actualización completada para {remitente}\n")

# ==== 3. Flujo principal ====
def main():
    ruta = "./GENERADOS/contactos_filtrados.xlsx"
    if not os.path.exists(ruta):
        print("❌ No se encontró el archivo Excel.")
        return

    # Cargar remitentes únicos desde columna "Correo"
    wb = load_workbook(ruta)
    ws = wb.active
    encabezados = [cell.value for cell in ws[1]]

    if "Correo" not in encabezados:
        print("❌ La columna 'Correo' no existe en el archivo.")
        return

    col_correo = encabezados.index("Correo") + 1
    remitentes_unicos = set()
    for fila in range(2, ws.max_row + 1):
        val = ws.cell(row=fila, column=col_correo).value
        if val:
            remitentes_unicos.add(str(val).strip().lower())

    # Ejecutar verificación para cada remitente
    for remitente in remitentes_unicos:
        perfil = next((p for p in perfiles.values() if p["correo"].lower() == remitente), None)
        if perfil:
            verificar_estado_por_remitente(ruta, perfil["correo"], perfil["app_password"])
        else:
            print(f"⚠️ No hay credenciales para el remitente: {remitente}")

if __name__ == "__main__":
    main()
