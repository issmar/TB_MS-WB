import os
import time
import urllib.parse
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def enviar_mensajes_wa(limite_envios):
    RUTA_EXCEL = "./GENERADOS/contactos_filtrados.xlsx"
    CHROMEDRIVER_PATH = "chromedriver.exe"
    PROFILE_PATH = "user-data-dir=C:/Users/ALPHA_PC/AppData/Local/Google/Chrome/User Data/WhatsappProfile"
    WA_URL = "https://web.whatsapp.com/"

    if not os.path.exists(RUTA_EXCEL):
        print("❌ No se encontró el archivo 'contactos_filtrados.xlsx' en ./GENERADOS/")
        return

    print(f"📄 Cargando archivo: {RUTA_EXCEL}")
    wb = load_workbook(RUTA_EXCEL)
    ws = wb.active
    headers = [cell.value.strip().lower().replace(" ", "").replace(".", "").replace("í", "i") for cell in ws[1]]

    try:
        col_tel = headers.index("telefono") + 1
        col_envio_wa = headers.index("estadoenviowa") + 1
        col_resp_wa = headers.index("estadorespuestawa") + 1
    except ValueError:
        print("❌ Faltan columnas requeridas: 'Teléfono', 'Estado Envío WA', 'Estado Respuesta WA'")
        return

    print("\n🚀 Iniciando WhatsApp Web...")
    options = Options()
    options.add_argument(PROFILE_PATH)

    try:
        service = Service(CHROMEDRIVER_PATH)
        driver = webdriver.Chrome(service=service, options=options)
        driver.get(WA_URL)

        print("📱 Esperando que se inicie sesión en WhatsApp Web...")
        WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.ID, "pane-side")))
    except Exception as e:
        print(f"❌ No se pudo iniciar WhatsApp Web: {e}")
        return

    mensaje = (
        "Te escribimos de NIFLOR.\n"
        "🚛 BASE EN ALTAMIRA, TAMPS.\n"
        "4 TRACTOCAMIONES FULL disponibles.\n"
        "¿Tienes cargas que no se están moviendo por falta de transporte?\n\n"
        "Somos personas trabajadoras y de confianza, con rutas activas a Nuevo León, San Luis Potosí, Jalisco, Coahuila y Tamaulipas.\n"
        "Empresa nueva, flotilla pequeña = atención personalizada, disponibilidad inmediata y seguimiento puntual.\n\n"
        "¿Te molesta si cotizamos alguno de tus requerimientos?\n\n"
        "Saludos,\nEnrique Delgado\n833 236 66 62"
    )
    mensaje_url = urllib.parse.quote(mensaje)

    enviados = 0

    for fila in range(2, ws.max_row + 1):
        if enviados >= limite_envios:
            print(f"\n🛑 Límite de envíos alcanzado: {limite_envios}")
            break

        celda_tel = ws.cell(row=fila, column=col_tel)
        celda_envio = ws.cell(row=fila, column=col_envio_wa)
        celda_resp = ws.cell(row=fila, column=col_resp_wa)

        estado_envio = str(celda_envio.value).strip().lower() if celda_envio.value else ""
        if estado_envio != "pendiente":
            continue

        telefono = str(celda_tel.value).strip() if celda_tel.value else ""
        if not telefono or not telefono.isdigit():
            celda_envio.value = "No Aplica"
            celda_resp.value = "No Aplica"
            print(f"❌ Fila {fila}: Teléfono vacío o inválido → No Aplica")
            wb.save(RUTA_EXCEL)
            continue

        telefono = ''.join(filter(str.isdigit, telefono))
        if not telefono.startswith("52"):
            telefono = "521" + telefono

        enlace = f"https://web.whatsapp.com/send?phone={telefono}&text={mensaje_url}"
        driver.get(enlace)
        time.sleep(5)

        try:
            boton_enviar = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@aria-label="Enviar"]'))
            )
            boton_enviar.click()
            celda_envio.value = "Enviado"
            celda_resp.value = "Pendiente"
            enviados += 1
            print(f"✅ Fila {fila}: Mensaje enviado a {telefono} ({enviados}/{limite_envios})")
        except Exception:
            celda_envio.value = "No Existe"
            celda_resp.value = "No Aplica"
            print(f"❌ Fila {fila}: No se pudo enviar mensaje a {telefono}")

        wb.save(RUTA_EXCEL)
        time.sleep(2)

    driver.quit()
    print(f"\n✅ Proceso finalizado. Total enviados: {enviados} / {limite_envios}")
    print("📄 Archivo actualizado: contactos_filtrados.xlsx")

if __name__ == "__main__":
    try:
        cantidad = int(input("🔢 ¿Cuántos mensajes de WhatsApp deseas enviar?: "))
        if cantidad <= 0:
            print("❌ Debes ingresar un número mayor a cero.")
        else:
            enviar_mensajes_wa(cantidad)
    except ValueError:
        print("❌ Entrada inválida. Ingresa un número entero.")
