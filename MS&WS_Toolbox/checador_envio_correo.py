import os
import imaplib
import email
import re
import datetime as dt
from email import policy
from openpyxl import load_workbook
from openpyxl.styles import Font

def verificar_rebotes_correo(usuario_email, app_password):
    carpeta = "./Generados"
    archivos = [f for f in os.listdir(carpeta) if f.endswith(".xlsx")]

    if not archivos:
        print("❌ No hay archivos .xlsx en la carpeta 'Generados'.")
        return

    print("\n📂 Archivos disponibles:")
    for idx, nombre in enumerate(archivos, 1):
        print(f"{idx}) {nombre}")

    try:
        opcion = int(input("\nSeleccione el número del archivo: "))
        archivo = archivos[opcion - 1]
    except (ValueError, IndexError):
        print("❌ Opción inválida.")
        return

    ruta_archivo = os.path.join(carpeta, archivo)
    print(f"\n🔎 Analizando: {archivo}")

    print("\n⏳ Conectando a Gmail IMAP y buscando rebotes…")
    try:
        imap = imaplib.IMAP4_SSL("imap.gmail.com", 993)
        imap.login(usuario_email, app_password)
    except imaplib.IMAP4.error as err:
        print(f"❌ Error de autenticación IMAP: {err}")
        return

    imap.select("INBOX")
    since_date = (dt.datetime.utcnow() - dt.timedelta(days=30)).strftime("%d-%b-%Y")

    status, data = imap.search(
        None,
        f'(FROM "mailer-daemon@googlemail.com" SINCE {since_date})'
    )

    if status != "OK":
        print("❌ No se pudo realizar la búsqueda IMAP.")
        imap.logout()
        return

    ids = data[0].split()
    bounced_set = set()
    regex_email = re.compile(r"[\w\.-]+@[\w\.-]+\.\w+")

    for msg_id in ids:
        status, msg_data = imap.fetch(msg_id, "(RFC822)")
        if status != "OK":
            continue

        msg = email.message_from_bytes(msg_data[0][1], policy=policy.default)
        body_text = ""

        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    body_text = part.get_content()
                    break
        else:
            body_text = msg.get_content()

        bounced_emails = regex_email.findall(body_text)
        for b in bounced_emails:
            bounced_set.add(b.lower())

    imap.logout()

    print(f"📛 Correos rebotados encontrados: {len(bounced_set)}")

    wb = load_workbook(ruta_archivo)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    try:
        col_email     = headers.index("E Mail") + 1
        col_envio     = headers.index("Estado Envío Correo") + 1
        col_respuesta = headers.index("Estado Respuesta Correo") + 1
    except ValueError:
        print("❌ El archivo no contiene las columnas requeridas.")
        return

    rojo = Font(color="FFFF0000")
    bounces_marked = 0

    for row in range(2, ws.max_row + 1):
        email_cell = ws.cell(row=row, column=col_email)
        envio_cell = ws.cell(row=row, column=col_envio)
        resp_cell  = ws.cell(row=row, column=col_respuesta)

        correo = str(email_cell.value or "").strip().lower()
        estado_envio = str(envio_cell.value or "").strip().lower()

        if estado_envio != "enviado":
            continue

        if correo in bounced_set:
            envio_cell.value = "Rebotado"
            resp_cell.value  = "Rebotado"
            envio_cell.font = rojo
            resp_cell.font = rojo
            bounces_marked += 1
            print(f"🚨 Rebotado detectado: {correo}")
            wb.save(ruta_archivo)  # Guardar por cada rebote encontrado

    print(f"\n✅ Archivo actualizado: {archivo}")
    print(f"🔴 Total filas marcadas como 'Rebotado': {bounces_marked}")
