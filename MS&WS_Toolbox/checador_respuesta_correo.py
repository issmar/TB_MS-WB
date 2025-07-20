import os
import imaplib
import email
import re
import datetime as dt
from email import policy
from openpyxl import load_workbook
from openpyxl.styles import Font

def verificar_respuestas_correo(user_email, app_password):
    FOLDER = "./Generados"
    files = [f for f in os.listdir(FOLDER) if f.endswith(".xlsx")]

    if not files:
        print("❌ No hay archivos .xlsx en la carpeta 'Generados'.")
        return

    print("\n📂 Archivos disponibles:")
    for idx, fname in enumerate(files, 1):
        print(f"{idx}) {fname}")

    try:
        choice = int(input("\nSeleccione el número del archivo: "))
        chosen_file = files[choice - 1]
    except (ValueError, IndexError):
        print("❌ Opción inválida.")
        return

    path_file = os.path.join(FOLDER, chosen_file)
    print(f"\n📬 Buscando respuestas en: {chosen_file}")

    try:
        imap = imaplib.IMAP4_SSL("imap.gmail.com", 993)
        imap.login(user_email, app_password)
        imap.select("INBOX")
    except Exception as e:
        print(f"❌ No se pudo conectar a IMAP: {e}")
        return

    since_date = (dt.datetime.utcnow() - dt.timedelta(days=30)).strftime("%d-%b-%Y")
    status, data = imap.search(None, f'(SINCE {since_date})')

    if status != "OK":
        print("❌ Error buscando respuestas.")
        imap.logout()
        return

    ids = data[0].split()
    respuestas_set = set()
    regex_email = re.compile(r"[\w\.-]+@[\w\.-]+\.\w+")

    for msg_id in ids:
        status, msg_data = imap.fetch(msg_id, "(RFC822)")
        if status != "OK":
            continue

        msg = email.message_from_bytes(msg_data[0][1], policy=policy.default)
        remitente = msg.get("From")

        if remitente:
            encontrado = re.findall(regex_email, remitente)
            if encontrado:
                respuestas_set.add(encontrado[0].lower())

    imap.logout()
    print(f"📬 Correos con respuesta detectados: {len(respuestas_set)}")

    wb = load_workbook(path_file)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    try:
        col_email     = headers.index("E Mail") + 1
        col_envio     = headers.index("Estado Envío Correo") + 1
        col_respuesta = headers.index("Estado Respuesta Correo") + 1
    except ValueError:
        print("❌ El archivo no contiene las columnas requeridas.")
        return

    verde = Font(color="FF008000")
    respuestas_marcadas = 0

    for row in range(2, ws.max_row + 1):
        email_cell = ws.cell(row=row, column=col_email)
        envio_cell = ws.cell(row=row, column=col_envio)
        resp_cell  = ws.cell(row=row, column=col_respuesta)

        correo = str(email_cell.value or "").strip().lower()
        estado_envio = str(envio_cell.value or "").strip().lower()

        if estado_envio != "enviado":
            continue

        if correo in respuestas_set:
            resp_cell.value = "Respondido"
            resp_cell.font = verde
            respuestas_marcadas += 1
            print(f"✅ Respuesta detectada: {correo}")
            wb.save(path_file)  # Guardar tras cada fila actualizada

    print(f"\n✅ Archivo actualizado: {chosen_file}")
    print(f"🟢 Total respuestas marcadas: {respuestas_marcadas}")
