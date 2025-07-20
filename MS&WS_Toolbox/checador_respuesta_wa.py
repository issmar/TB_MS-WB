import os
import time
from openpyxl import load_workbook
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException

def verificar_respuestas_wa():
    CHROME_PROFILE_PATH = "C:/Users/ALPHA_PC/AppData/Local/Google/Chrome/User Data/WhatsappProfile"
    CHROMEDRIVER_PATH = "chromedriver.exe"
    FOLDER = "./Generados"

    def iniciar_whatsapp():
        options = Options()
        options.add_argument(f"user-data-dir={CHROME_PROFILE_PATH}")
        service = Service(CHROMEDRIVER_PATH)
        driver = webdriver.Chrome(service=service, options=options)
        driver.get("https://web.whatsapp.com")

        try:
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((By.ID, "pane-side"))
            )
            print("✅ Sesión de WhatsApp iniciada.")
        except TimeoutException:
            print("❌ No se pudo iniciar sesión en WhatsApp.")
            driver.quit()
            return None
        return driver

    archivos = [f for f in os.listdir(FOLDER) if f.endswith(".xlsx")]
    if not archivos:
        print("❌ No hay archivos .xlsx en la carpeta 'Generados'.")
        return

    print("\n📂 Archivos disponibles:")
    for i, fname in enumerate(archivos, 1):
        print(f"{i}) {fname}")

    try:
        idx = int(input("\nSeleccione el número del archivo: "))
        selected_file = archivos[idx - 1]
    except:
        print("❌ Opción inválida.")
        return

    path = os.path.join(FOLDER, selected_file)
    print(f"\n🔍 Abriendo archivo: {selected_file}")

    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    try:
        col_tel = next(i + 1 for i, v in enumerate(headers) if str(v).strip().lower().startswith("telefono"))
        col_envio = headers.index("Estado Envío WA") + 1
        col_respuesta = headers.index("Estado Respuesta WA") + 1
    except ValueError:
        print("❌ Columnas requeridas no encontradas.")
        return

    driver = iniciar_whatsapp()
    if not driver:
        return

    verde = Font(color="FF008000")

    for row in range(2, ws.max_row + 1):
        tel_cell = ws.cell(row=row, column=col_tel)
        envio_cell = ws.cell(row=row, column=col_envio)
        respuesta_cell = ws.cell(row=row, column=col_respuesta)

        tel = str(tel_cell.value or "").strip()
        estado_envio = str(envio_cell.value or "").strip().lower()
        estado_respuesta = str(respuesta_cell.value or "").strip().lower()

        if estado_envio != "enviado" or estado_respuesta == "respondido":
            continue

        if not tel:
            continue

        tel = ''.join(filter(str.isdigit, tel))
        if not tel:
            continue

        link = f"https://web.whatsapp.com/send?phone=521{tel}"
        print(f"📱 Revisando: {tel}")
        driver.get(link)
        time.sleep(5)

        try:
            msg_xpath = '(//div[contains(@class,"message-in")]//div[@class="copyable-text"])[last()]'
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, msg_xpath))
            )
            respuesta_cell.value = "Respondido"
            respuesta_cell.font = verde
            wb.save(path)  # Guardar fila con respuesta
            print(f"✅ Respondió: {tel}")
        except:
            print(f"⏳ Sin respuesta: {tel}")

        time.sleep(2)

    driver.quit()
    print(f"\n✅ Archivo actualizado: {selected_file}")
