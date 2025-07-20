import os
import socket
import shutil
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Color

def ejecutar_checador_dns():
    # Configurar colores
    verde = "FF008000"
    rojo = "FFFF0000"

    # Carpeta donde están los archivos generados
    carpeta_generados = "./Generados"
    archivos_excel = [f for f in os.listdir(carpeta_generados) if f.endswith(".xlsx")]

    if not archivos_excel:
        print("❌ No se encontraron archivos .xlsx en la carpeta 'Generados'.")
        return

    # Mostrar archivos disponibles
    print("\n📂 Archivos disponibles:")
    for idx, archivo in enumerate(archivos_excel, start=1):
        print(f"{idx}) {archivo}")

    # Elegir archivo
    try:
        opcion = int(input("\nSeleccione el número del archivo que desea procesar: "))
        if opcion < 1 or opcion > len(archivos_excel):
            print("❌ Número fuera de rango.")
            return
    except ValueError:
        print("❌ Entrada inválida.")
        return

    archivo_seleccionado = archivos_excel[opcion - 1]
    ruta_archivo = os.path.join(carpeta_generados, archivo_seleccionado)

    # Crear respaldo del archivo original
    respaldo_path = ruta_archivo.replace(".xlsx", "_respaldo.xlsx")
    shutil.copy2(ruta_archivo, respaldo_path)
    print(f"📄 Respaldo creado: {os.path.basename(respaldo_path)}")

    # Cargar DataFrame para validar columnas
    df = pd.read_excel(ruta_archivo)

    if "E Mail" not in df.columns or "Estado Dominio" not in df.columns:
        print("❌ El archivo no contiene las columnas necesarias: 'E Mail' y 'Estado Dominio'")
        return

    # Cargar archivo con openpyxl para edición con formato
    wb = load_workbook(ruta_archivo)
    ws = wb.active

    # Identificar índices de columna
    cabeceras = [cell.value for cell in ws[1]]
    col_email = cabeceras.index("E Mail") + 1
    col_estado = cabeceras.index("Estado Dominio") + 1

    # Procesar fila por fila
    print(f"\n🔍 Verificando dominios en: {archivo_seleccionado}\n")
    for row in range(2, ws.max_row + 1):
        email = ws.cell(row=row, column=col_email).value
        if not email or str(email).strip() == "":
            for col_name in ["Estado Dominio", "Estado Envío Correo", "Estado Respuesta Correo"]:
                if col_name in cabeceras:
                    col_index = cabeceras.index(col_name) + 1
                    ws.cell(row=row, column=col_index).value = "N/A"
                    ws.cell(row=row, column=col_index).font = Font(color=Color(rgb=rojo))
            print(f"⚠️ Fila {row}: E-mail vacío, columnas marcadas como 'N/A'")
            wb.save(ruta_archivo)
            continue

        dominio = str(email).strip().split("@")[-1].lower()
        try:
            socket.gethostbyname(dominio)
            ws.cell(row=row, column=col_estado).value = "Existe"
            ws.cell(row=row, column=col_estado).font = Font(color=Color(rgb=verde))
            print(f"✅ {dominio}: Existe")
        except socket.gaierror:
            ws.cell(row=row, column=col_estado).value = "No existe"
            ws.cell(row=row, column=col_estado).font = Font(color=Color(rgb=rojo))
            print(f"❌ {dominio}: No existe")

        # Guardar después de procesar cada fila
        wb.save(ruta_archivo)

    print(f"\n✅ Verificación completada y archivo actualizado: {archivo_seleccionado}")

# Ejecutar la función
if __name__ == "__main__":
    ejecutar_checador_dns()
