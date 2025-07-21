import os
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import urllib.parse

def enviar_mensajes_wa():
    def normalizar(texto):
        return str(texto).strip().lower().replace(" ", "").replace(".", "").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")

    FOLDER = "./Generados"
    CHROMEDRIVER_PATH = "chromedriver.exe"
    PROFILE_PATH = "user-data-dir=C:/Users/angel/AppData/Local/Google/Chrome/User Data/WhatsappProfile"
    WA_URL = "https://web.whatsapp.com/"

    verde = Font(color="FF008000")
    rojo = Font(color="FFFF0000")
    naranja = Font(color="FFFF9900")

    archivos = [f for f in os.listdir(FOLDER) if f.endswith(".xlsx")]

    if not archivos:
        print("❌ No se encontraron archivos .xlsx en la carpeta 'Generados'.")
        return

    print("📂 Archivos disponibles:")
    for idx, nombre in enumerate(archivos, 1):
        print(f"{idx}) {nombre}")

    try:
        opcion = int(input("\nSeleccione el número del archivo: "))
        archivo = archivos[opcion - 1]
    except:
        print("❌ Opción inválida.")
        return

    ruta = os.path.join(FOLDER, archivo)
    print(f"\n📄 Abriendo: {archivo}")

    wb = load_workbook(ruta)
    ws = wb.active
    headers = [normalizar(cell.value) for cell in ws[1]]

    try:
        col_tel = headers.index("telefono") + 1
        col_envio_wa = headers.index("estadoenviowa") + 1
        col_resp_wa = headers.index("estadorespuestawa") + 1
    except ValueError:
        print("❌ Faltan columnas requeridas: 'Teléfono', 'Estado Envío WA', 'Estado Respuesta WA'")
        print(f"🧩 Encabezados detectados: {headers}")
        return

    print("\n🚀 Iniciando WhatsApp Web...")
    options = Options()
    options.add_argument(PROFILE_PATH)

    try:
        service = Service(CHROMEDRIVER_PATH)
        driver = webdriver.Chrome(service=service, options=options)
        driver.get(WA_URL)

        print("📱 Escanea el código QR si aún no has iniciado sesión...")
        WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.ID, "pane-side")))
    except Exception as e:
        print(f"❌ No se pudo iniciar WhatsApp Web: {e}")
        return

    mensaje = (
        "Te escribimos de NIFLOR.\n"
        "🚛 BASE EN ALTAMIRA, TAMPS.\n"
        "4 TRACTOCAMIONES FULL disponibles.\n"
        "¿Tienes cargas que no se están moviendo por falta de transporte?\n\n"
        "Somos personas trabajadoras y de confianza, con rutas activas a Nuevo León, San Luis Potosí, Jalisco, Coahuila y Tamaulipas.\n"
        "Empresa nueva, flotilla pequeña = atención personalizada, disponibilidad inmediata y seguimiento puntual.\n\n"
        "¿Te molesta si cotizamos alguno de tus requerimientos?\n\n"
        "Saludos,\nEnrique Delgado\n833 236 66 62"
    )
    mensaje_url = urllib.parse.quote(mensaje)

    for fila in range(2, ws.max_row + 1):
        celda_tel = ws.cell(row=fila, column=col_tel)
        celda_envio = ws.cell(row=fila, column=col_envio_wa)
        celda_resp = ws.cell(row=fila, column=col_resp_wa)

        telefono = str(celda_tel.value).strip() if celda_tel.value else ""
        if not telefono:
            celda_envio.value = "N/A"
            celda_envio.font = rojo
            celda_resp.value = "N/A"
            celda_resp.font = rojo
            wb.save(ruta)
            continue

        telefono = ''.join(filter(str.isdigit, telefono))
        if not telefono.startswith("52"):
            telefono = "521" + telefono

        enlace = f"https://web.whatsapp.com/send?phone={telefono}&text={mensaje_url}"
        driver.get(enlace)
        time.sleep(5)

        try:
            boton_enviar = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@aria-label="Enviar"]'))
            )
            boton_enviar.click()
            celda_envio.value = "Enviado"
            celda_envio.font = verde
            if celda_resp.value in [None, "", "Pendiente"]:
                celda_resp.value = "Pendiente"
                celda_resp.font = naranja
            time.sleep(2)
            wb.save(ruta)  # Guardar tras envío exitoso
        except Exception:
            celda_envio.value = "No existe"
            celda_envio.font = rojo
            celda_resp.value = "N/A"
            celda_resp.font = rojo
            wb.save(ruta)  # Guardar tras fallo
            continue

    driver.quit()
    print(f"\n✅ Archivo actualizado: {archivo}")
