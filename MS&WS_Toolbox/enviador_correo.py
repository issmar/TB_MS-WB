import os
import smtplib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

def enviar_correos(usuario_email, usuario_password):
    carpeta = "./Generados"
    pdf_filename = "NIFLOR.pdf"
    pdf_path = os.path.abspath(pdf_filename)

    if not os.path.isfile(pdf_path):
        print(f"❌ El archivo PDF '{pdf_filename}' no fue encontrado.")
        return

    archivos = [f for f in os.listdir(carpeta) if f.endswith(".xlsx")]

    if not archivos:
        print("❌ No se encontraron archivos en la carpeta 'Generados'.")
        return

    print("\n📂 Archivos disponibles en 'Generados':")
    for i, archivo in enumerate(archivos, start=1):
        print(f"{i}) {archivo}")

    try:
        opcion = int(input("\nSeleccione el número del archivo a procesar: "))
        archivo_seleccionado = archivos[opcion - 1]
    except:
        print("❌ Opción inválida.")
        return

    ruta_archivo = os.path.join(carpeta, archivo_seleccionado)
    print(f"\n📤 Iniciando envío de correos en: {archivo_seleccionado}\n")

    df = pd.read_excel(ruta_archivo)

    if "E Mail" not in df.columns or "Estado Dominio" not in df.columns or "Estado Envío Correo" not in df.columns:
        print("❌ El archivo debe contener las columnas necesarias.")
        return

    wb = load_workbook(ruta_archivo)
    ws = wb.active

    header = [cell.value for cell in ws[1]]
    col_email = header.index("E Mail") + 1
    col_dominio = header.index("Estado Dominio") + 1
    col_envio = header.index("Estado Envío Correo") + 1

    verde = Font(color="FF008000")
    rojo = Font(color="FFFF0000")
    enviados_set = set()

    for fila in range(2, ws.max_row + 1):
        email_cell = ws.cell(row=fila, column=col_email)
        dominio_cell = ws.cell(row=fila, column=col_dominio)
        envio_cell = ws.cell(row=fila, column=col_envio)

        correo = str(email_cell.value).strip().lower() if email_cell.value else ""
        dominio_estado = str(dominio_cell.value).strip().lower() if dominio_cell.value else ""
        estado_actual = str(envio_cell.value).strip().lower() if envio_cell.value else ""

        if not correo:
            envio_cell.value = "N/A"
            envio_cell.font = rojo
            print(f"⚠️ Correo vacío en fila {fila} → marcado como 'N/A'")
            wb.save(ruta_archivo)
            continue

        if dominio_estado != "existe":
            envio_cell.value = "N/A"
            envio_cell.font = rojo
            print(f"❌ Dominio no válido: {correo}")
            wb.save(ruta_archivo)
            continue

        if correo in enviados_set:
            envio_cell.value = "Repetido"
            envio_cell.font = rojo
            print(f"🔁 Repetido: {correo}")
            wb.save(ruta_archivo)
            continue

        try:
            mensaje = MIMEMultipart()
            mensaje["From"] = usuario_email
            mensaje["To"] = correo
            mensaje["Subject"] = "NIFLOR: Servicio de confianza / 4 tractos servicio full"

            cuerpo_html = """
            <b>🚛 BASE EN ALTAMIRA, TAMPS.</b><br>
            <b>4 TRACTOCAMIONES CON SERVICIO FULL, DISPONIBLES</b><br><br>
            ¿Tienes cargas que no se están moviendo por falta de transporte?<br><br>
            Formada por personas trabajadoras y de confianza, con rutas activas a <b>Nuevo León, San Luis Potosí, Jalisco, Coahuila y Tamaulipas</b>.<br><br>
            Somos una empresa nueva y al tener una flotilla pequeña, podemos darte <b>atención personalizada, disponibilidad inmediata y seguimiento puntual</b>.<br><br>
            ¿Te molesta si cotizamos alguno de tus requerimientos?<br><br>
            Saludos,<br>
            <b>Enrique Delgado</b><br>
            <b>833 236 66 62</b>
            """

            cuerpo = MIMEText(cuerpo_html, "html")
            mensaje.attach(cuerpo)

            # Adjuntar el PDF
            with open(pdf_path, "rb") as archivo_pdf:
                parte_pdf = MIMEApplication(archivo_pdf.read(), _subtype="pdf")
                parte_pdf.add_header("Content-Disposition", "attachment", filename=pdf_filename)
                mensaje.attach(parte_pdf)

            # Enviar correo
            with smtplib.SMTP("smtp.gmail.com", 587) as server:
                server.starttls()
                server.login(usuario_email, usuario_password)
                server.sendmail(usuario_email, correo, mensaje.as_string())

            envio_cell.value = "Enviado"
            envio_cell.font = verde
            enviados_set.add(correo)
            print(f"✅ Enviado: {correo}")
            wb.save(ruta_archivo)

        except Exception as e:
            envio_cell.value = "N/A"
            envio_cell.font = rojo
            print(f"❌ Error al enviar a {correo}: {e}")
            wb.save(ruta_archivo)

    print(f"\n✅ Archivo actualizado: {archivo_seleccionado}")
    print(f"📬 Total de correos enviados: {len(enviados_set)}")
