# separador_filas.py

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

def ejecutar_separador():
    archivo_principal = "lista_contactos.xlsx"
    ruta = os.path.join(".", archivo_principal)

    if not os.path.exists(ruta):
        print("❌ No se encontró 'lista_contactos.xlsx'")
        return

    # Cargar todas las hojas
    xls = pd.ExcelFile(ruta)
    hojas = xls.sheet_names

    print("\n📄 Hojas disponibles:")
    for i, hoja in enumerate(hojas, 1):
        print(f"{i}) {hoja}")

    try:
        opcion = int(input("Seleccione el número de hoja a procesar: "))
        hoja_seleccionada = hojas[opcion - 1]
    except (ValueError, IndexError):
        print("❌ Selección inválida.")
        return

    df = xls.parse(hoja_seleccionada)
    total_filas = len(df)

    print(f"\n✅ Total de filas en '{hoja_seleccionada}': {total_filas}")
    inicio = int(input("🔢 Desde qué fila desea empezar (1-indexado): "))
    fin = int(input("🔢 Hasta qué fila desea usar (inclusive): "))

    if inicio < 1 or fin > total_filas or inicio > fin:
        print("❌ Rango inválido.")
        return

    df_corte = df.iloc[inicio - 1:fin].copy()

    # Agregar columnas con estado en blanco
    nuevas_columnas = [
        "Estado Dominio",
        "Estado Envío Correo",
        "Estado Respuesta Correo",
        "Estado Envío WA",
        "Estado Respuesta WA"
    ]

    for col in nuevas_columnas:
        df_corte[col] = ""

    # Crear nombre del archivo de salida
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"{hoja_seleccionada.replace(' ', '')}_{inicio}a{fin}_{timestamp}.xlsx"
    ruta_salida = os.path.join("Generados", nombre_archivo)

    # Crear carpeta si no existe
    os.makedirs("Generados", exist_ok=True)

    # Guardar en Excel
    df_corte.to_excel(ruta_salida, index=False)

    # Estilizar nuevas columnas (color naranja)
    wb = load_workbook(ruta_salida)
    ws = wb.active
    encabezados = [cell.value for cell in ws[1]]
    naranja = Font(color="FFFF9900")

    for col_name in nuevas_columnas:
        if col_name in encabezados:
            col_idx = encabezados.index(col_name) + 1
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col_idx).font = naranja

    wb.save(ruta_salida)

    print(f"\n✅ Archivo generado: {ruta_salida}")
