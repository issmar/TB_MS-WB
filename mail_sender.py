import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import pandas as pd
import time
import imaplib
import email
import os
import socket
from urllib.parse import urlparse
from openpyxl import load_workbook
from openpyxl.styles import Font, Color
import re

# === 1. Configuración global ===
smtp_server = "smtp.gmail.com"
smtp_port = 587
imap_server = "imap.gmail.com"
tu_correo = "luisrjm53@gmail.com"
tu_password = "pcloiobvfagkgfax" 
ruta_excel = "correos.xlsx"
nombre_archivo_pdf = "ejemplo.pdf"

# === 2. Función para manejar el Excel (leer y guardar) ===
def cargar_o_crear_df():
    """Carga el DataFrame desde el Excel o crea uno nuevo si no existe."""
    try:
        # Solución al problema de N/A y NaN
        df = pd.read_excel(ruta_excel, keep_default_na=False)
        print("💾 Archivo 'correos.xlsx' cargado con éxito.")
    except FileNotFoundError:
        print(f"❌ Error: No se encontró el archivo '{ruta_excel}'. Creando uno nuevo...")
        columnas = ["E Mail", "Nombre", "Sitio web", "Estatus Correo", "Estatus Envío", "Estatus Dominios", "Estatus Respuesta"]
        df = pd.DataFrame(columns=columnas)
        # Inicializar todos los estatus a "Pendiente"
        df["Estatus Correo"] = "Pendiente"
        df["Estatus Envío"] = "Pendiente"
        df["Estatus Dominios"] = "Pendiente"
        df["Estatus Respuesta"] = "Pendiente"
        
    # Unificar la columna 'Estatus' a 'Estatus Correo' si existe
    if "Estatus" in df.columns and "Estatus Correo" not in df.columns:
        df.rename(columns={"Estatus": "Estatus Correo"}, inplace=True)
        
    # Asegurarse de que existan las columnas necesarias y que sean de tipo string
    for col in ["Estatus Correo", "Estatus Envío", "Estatus Dominios", "Estatus Respuesta"]:
        if col not in df.columns:
            df[col] = "Pendiente"
        df[col] = df[col].astype(str)

    return df

def guardar_df_con_formato(df):
    """
    Guarda el DataFrame en el archivo de Excel y aplica los formatos de color
    según el estado de cada columna.
    """
    print("\n💾 Guardando y aplicando formato al archivo de Excel...")
    
    try:
        with pd.ExcelWriter(ruta_excel, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            workbook = writer.book
            sheet = writer.sheets['Sheet1']

            columnas_a_colorear = ["Sitio web", "E Mail", "Estatus Correo", "Estatus Envío", "Estatus Dominios", "Estatus Respuesta"]
            indices = {col: -1 for col in columnas_a_colorear}
            for idx, cell in enumerate(sheet[1]):
                if cell.value in indices:
                    indices[cell.value] = idx + 1
            
            verde = "FF008000"  
            rojo = "FFFF0000"
            naranja = "FFFF8C00" 
            
            for index, row in df.iterrows():
                fila_excel = index + 2
                
                # Colorear en base a los estatus
                estatus_dom = str(row["Estatus Dominios"]).strip()
                estatus_correo = str(row["Estatus Correo"]).strip()
                estatus_envio = str(row["Estatus Envío"]).strip()
                estatus_respuesta = str(row["Estatus Respuesta"]).strip()
                
                # Regla de color para Estatus Dominios y Sitio web
                if estatus_dom == "Existe":
                    if indices["Sitio web"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Sitio web"]).font = Font(color=Color(rgb=verde))
                    if indices["Estatus Dominios"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Estatus Dominios"]).font = Font(color=Color(rgb=verde))
                elif estatus_dom == "No existe":
                    if indices["Sitio web"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Sitio web"]).font = Font(color=Color(rgb=rojo))
                    if indices["Estatus Dominios"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Estatus Dominios"]).font = Font(color=Color(rgb=rojo))
                else: # Pendiente
                    if indices["Estatus Dominios"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Estatus Dominios"]).font = Font(color=Color(rgb=naranja))

                # Regla de color para Estatus Correo
                if estatus_correo == "Enviado":
                    if indices["Estatus Correo"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Estatus Correo"]).font = Font(color=Color(rgb=verde))
                elif estatus_correo == "N/A":
                    if indices["E Mail"] != -1:
                        sheet.cell(row=fila_excel, column=indices["E Mail"]).font = Font(color=Color(rgb=rojo))
                    if indices["Estatus Correo"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Estatus Correo"]).font = Font(color=Color(rgb=rojo))
                elif estatus_correo == "Pendiente":
                    if indices["Estatus Correo"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Estatus Correo"]).font = Font(color=Color(rgb=naranja))

                # Regla de color para Estatus Envío
                if estatus_envio == "Confirmado":
                    if indices["Estatus Envío"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Estatus Envío"]).font = Font(color=Color(rgb=verde))
                elif estatus_envio == "Rechazado":
                    if indices["Estatus Envío"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Estatus Envío"]).font = Font(color=Color(rgb=rojo))
                elif estatus_envio in ["En Espera", "Pendiente"]:
                    if indices["Estatus Envío"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Estatus Envío"]).font = Font(color=Color(rgb=naranja))
                elif estatus_envio == "N/A":
                    if indices["Estatus Envío"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Estatus Envío"]).font = Font(color=Color(rgb=rojo))
                
                # Regla de color para Estatus Respuesta
                if estatus_respuesta == "Respuesta":
                    if indices["Estatus Respuesta"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Estatus Respuesta"]).font = Font(color=Color(rgb=verde))
                elif estatus_respuesta in ["En Espera", "Pendiente"]:
                    if indices["Estatus Respuesta"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Estatus Respuesta"]).font = Font(color=Color(rgb=naranja))
                elif estatus_respuesta == "N/A":
                    if indices["Estatus Respuesta"] != -1:
                        sheet.cell(row=fila_excel, column=indices["Estatus Respuesta"]).font = Font(color=Color(rgb=rojo))

            workbook.save(ruta_excel)
        print("✅ Archivo actualizado con éxito, manteniendo el formato.")
    except Exception as e:
        print(f"❌ Error al guardar el archivo: {e}. Asegúrate de que no esté abierto.")

# === 3. Funciones de validación ===
def validar_correo_sintaxis(correo):
    """Verifica la sintaxis de un correo electrónico usando una expresión regular."""
    regex = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    return re.match(regex, str(correo))

def es_dominio_valido_para_envio(row):
    """
    Determina si un correo puede ser enviado basándose en el estatus del dominio.
    Solo bloquea si el dominio del sitio web no existe Y el dominio del correo coincide.
    """
    correo = str(row["E Mail"]).strip().lower()
    sitio_web = str(row.get("Sitio web", "")).strip().lower()
    estatus_dom = str(row.get("Estatus Dominios", "")).strip()
    
    if estatus_dom == "No existe":
        dominio_correo = correo.split("@")[-1] if "@" in correo else ""
        dominio_sitio = sitio_web.replace("www.", "").replace("http://", "").replace("https://", "").split("/")[0]
        return dominio_correo != dominio_sitio
    return True

# === 4. Función para enviar correos con adjunto ===
def enviar_correos(df, limit):
    """Envía un número limitado de correos a los destinatarios con estatus 'Pendiente'."""
    correos_a_enviar = df[
        (df["Estatus Correo"].str.lower() == "pendiente") & 
        df["E Mail"].apply(validar_correo_sintaxis).astype(bool)
    ].head(limit).copy() # Se usa .head(limit) para limitar la cantidad a procesar

    if correos_a_enviar.empty:
        print("\n✅ No hay correos pendientes para enviar.")
        return df, 0

    ruta_pdf = os.path.join(os.getcwd(), nombre_archivo_pdf)
    if not os.path.exists(ruta_pdf):
        print(f"\n❌ Error: El archivo '{nombre_archivo_pdf}' no se encontró.")
        return df, 0

    print(f"\n--- Iniciando envío de {len(correos_a_enviar)} correo(s) ---")
    
    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(tu_correo, tu_password)

            for index, row in correos_a_enviar.iterrows():
                correo_destino = str(row["E Mail"]).strip().lower()
                
                if not es_dominio_valido_para_envio(row):
                    print(f"🚫 Salteado: {correo_destino} (Dominio del correo coincide con sitio web no existente)")
                    df.at[index, "Estatus Correo"] = "N/A"
                    df.at[index, "Estatus Envío"] = "N/A"
                    df.at[index, "Estatus Respuesta"] = "N/A"
                else:
                    msg = MIMEMultipart()
                    msg['From'] = tu_correo
                    msg['To'] = correo_destino
                    msg['Subject'] = "Prueba de envío con adjunto"
                    msg.attach(MIMEText("Hola, te envío un archivo de prueba adjunto.", 'plain'))
                    
                    with open(ruta_pdf, "rb") as attachment:
                        parte_adjunta = MIMEBase("application", "octet-stream")
                        parte_adjunta.set_payload(attachment.read())
                        encoders.encode_base64(parte_adjunta)
                        parte_adjunta.add_header("Content-Disposition", f"attachment; filename= {nombre_archivo_pdf}")
                        msg.attach(parte_adjunta)

                    server.sendmail(tu_correo, correo_destino, msg.as_string())
                    print(f"📤 Enviado a: {correo_destino}")
                    
                    df.at[index, "Estatus Correo"] = "Enviado"
                    df.at[index, "Estatus Envío"] = "En Espera"
                    df.at[index, "Estatus Respuesta"] = "En Espera"
                    time.sleep(1) 
        
        print("\n✅ Este lote de correos se procesó.")
        
    except smtplib.SMTPAuthenticationError:
        print("\n❌ Error de autenticación. Revisa tu correo y contraseña de aplicación.")
        return df, 0
    except Exception as e:
        print(f"\n❌ Ocurrió un error al enviar los correos: {e}")
        return df, 0
        
    return df, len(correos_a_enviar)

# === 5. Función para verificar correos rebotados ===
def verificar_rebotados(df, limit):
    """Verifica un número limitado de correos rebotados en las carpetas del IMAP."""
    correos_a_verificar = df[df["Estatus Envío"] == "En Espera"].head(limit)
    if correos_a_verificar.empty:
        print("\n✅ No hay correos para verificar si fueron rebotados.")
        return df, 0

    print("\n--- Verificando correos rebotados ---")
    
    rebotados = []
    try:
        with imaplib.IMAP4_SSL(imap_server) as mail:
            mail.login(tu_correo, tu_password)
            carpetas_a_buscar = ["INBOX", "[Gmail]/Spam", "[Gmail]/Trash"]

            for carpeta in carpetas_a_buscar:
                try:
                    mail.select(carpeta)
                    typ, data = mail.search(None, '(FROM "mailer-daemon")')
                    for num in data[0].split():
                        typ, msg_data = mail.fetch(num, '(RFC822)')
                        raw_email = msg_data[0][1]
                        msg = email.message_from_bytes(raw_email)
                        
                        body = ""
                        if msg.is_multipart():
                            for part in msg.walk():
                                if part.get_content_type() == "text/plain":
                                    body += part.get_payload(decode=True).decode(errors="ignore")
                        else:
                            body = msg.get_payload(decode=True).decode(errors="ignore")
                            
                        found = re.findall(r'[\w\.-]+@[\w\.-]+', body)
                        for correo in found:
                            if correo.lower() in correos_a_verificar["E Mail"].str.lower().values:
                                rebotados.append(correo.lower())
                except Exception as e:
                    print(f"⚠️  No se pudo revisar la carpeta '{carpeta}': {e}")
    except Exception as e:
        print(f"\n❌ Error al intentar conectar con IMAP: {e}")
        return df, 0

    rebotados = list(set(rebotados))
    
    correos_revisados_en_lote = 0
    for index, row in correos_a_verificar.iterrows():
        correo = row["E Mail"].strip().lower()
        if correo in rebotados:
            df.at[index, "Estatus Envío"] = "Rechazado"
        else:
            df.at[index, "Estatus Envío"] = "Confirmado"
        correos_revisados_en_lote += 1

    print("\n--- Verificación de rebotes de este lote completada ---")
    return df, correos_revisados_en_lote

# === 6. Función para verificar la existencia de sitios web ===
def verificar_sitios_web(df, limit):
    """Verifica la existencia de un número limitado de sitios web y actualiza el DataFrame."""
    print("\n--- Verificando existencia de sitios web (DNS) ---")

    # Obtener el subconjunto de filas a procesar
    rows_to_process = df[df["Estatus Dominios"] == "Pendiente"].head(limit)
    if rows_to_process.empty:
        print("✅ No hay sitios web pendientes para verificar.")
        return df, 0

    items_procesados = 0
    for index, row in rows_to_process.iterrows():
        url_str = str(row.get("Sitio web", "")).strip()
        
        existe_dns = False
        if not url_str or url_str.lower() in ['nan', 'nan.0']:
            df.at[index, "Estatus Dominios"] = "No existe"
            existe_dns = False
        else:
            if not url_str.startswith(('http://', 'https://')):
                url_str = 'http://' + url_str
            
            try:
                domain = urlparse(url_str).netloc
                if domain.startswith('www.'):
                    domain = domain[4:]
            except Exception:
                domain = None
            
            if domain:
                try:
                    socket.gethostbyname(domain)
                    existe_dns = True
                except (socket.gaierror, socket.error):
                    existe_dns = False
            
            resultado_texto = "Existe" if existe_dns else "No existe"
            df.at[index, "Estatus Dominios"] = resultado_texto
            print(f"{'✅' if existe_dns else '❌'} Dominio: {url_str} - {resultado_texto}")
        items_procesados += 1
    
    print("\n--- Verificación de este lote de sitios web completada ---")
    return df, items_procesados

# === 7. Función para verificar respuestas desde el Excel ===
def verificar_respuestas_desde_excel(df, limit):
    """Verifica un número limitado de respuestas de los correos y actualiza el DataFrame."""
    print("\n--- Iniciando verificación de respuestas desde el archivo de Excel ---")
    
    # Obtener la lista de correos únicos a verificar en este lote
    lista_correos_a_verificar = df[df["Estatus Respuesta"] == "En Espera"]["E Mail"].astype(str).str.strip().str.lower().head(limit).unique()
    
    if lista_correos_a_verificar.size == 0:
        print("\n✅ La columna 'E Mail' no tiene correos en 'En Espera'.")
        return df, 0

    respuestas_encontradas = 0
    try:
        with imaplib.IMAP4_SSL(imap_server) as mail:
            mail.login(tu_correo, tu_password)
            mail.select("INBOX")
            
            for correo in lista_correos_a_verificar:
                search_query = f'(FROM "{correo}")'
                typ, data = mail.search(None, search_query)
                
                if data[0]:
                    print(f"✅ Se encontró una respuesta de: {correo}")
                    # Buscar la fila en el DataFrame y actualizar el estatus
                    fila = df[df["E Mail"].str.lower() == correo]
                    if not fila.empty:
                        df.loc[fila.index, "Estatus Respuesta"] = "Respuesta"
                    respuestas_encontradas += 1
                else:
                    print(f"❌ No se encontró respuesta de: {correo}")
                    
    except Exception as e:
        print(f"\n❌ Ocurrió un error al conectar o buscar respuestas: {e}")
        return df, 0
        
    print("\n--- Verificación de respuestas de este lote completada ---")
    return df, respuestas_encontradas

# === 8. Menú principal ===
def main():
    """Menú principal de la aplicación."""
    df = cargar_o_crear_df()
    if df is None:
        return

    while True:
        os.system('cls' if os.name == 'nt' else 'clear')
        print("===================================")
        print("  HERRAMIENTA DE GESTIÓN DE CORREOS  ")
        print("===================================")
        print("1. Verificar existencia de sitios web (individual)")
        print("2. Verificar sitios web y enviar correos (Proceso automático)")
        print("3. Verificar correos rebotados")
        print("4. Verificar respuestas")
        print("5. Salir")
        print("-----------------------------------")
        
        opcion = input("Elige una opción (1-5): ")
        
        if opcion == "1":
            print("\n--- Iniciando proceso de verificación de sitios web (individual) ---")
            for i in range(3):
                df, items_procesados = verificar_sitios_web(df, 25)
                guardar_df_con_formato(df)
                if items_procesados == 0:
                    print("\n⚠️  No hay más sitios web pendientes para verificar. Proceso detenido.")
                    break
                if i < 2:
                    print("\n--- Esperando 5 segundos para el siguiente lote... ---")
                    # Para el uso real, cambiar time.sleep(5) por time.sleep(300)
                    time.sleep(5)
            print("\n✅ Proceso de verificación de sitios web completado.")
        elif opcion == "2":
            print("\n--- Iniciando proceso automático de verificación y envío ---")
            for i in range(3):
                df, sitios_procesados = verificar_sitios_web(df, 5)
                df, correos_enviados = enviar_correos(df, 5)
                guardar_df_con_formato(df)
                if sitios_procesados == 0 and correos_enviados == 0:
                    print("\n⚠️  No hay más ítems pendientes. Proceso detenido.")
                    break
                if i < 2:
                    print("\n--- Esperando 5 segundos para el siguiente lote... ---")
                    time.sleep(5)
            print("\n✅ Proceso automático completado.")
        elif opcion == "3":
            print("\n--- Iniciando proceso de verificación de rebotes ---")
            for i in range(3):
                df, correos_revisados = verificar_rebotados(df, 5)
                guardar_df_con_formato(df)
                if correos_revisados == 0:
                    print("\n⚠️  No hay más correos en espera de verificación. Proceso detenido.")
                    break
                if i < 2:
                    print("\n--- Esperando 5 segundos para el siguiente lote... ---")
                    time.sleep(5)
            print("\n✅ Proceso de verificación de rebotes completado.")
        elif opcion == "4":
            print("\n--- Iniciando proceso de verificación de respuestas ---")
            for i in range(3):
                df, respuestas_revisadas = verificar_respuestas_desde_excel(df, 5)
                guardar_df_con_formato(df)
                if respuestas_revisadas == 0:
                    print("\n⚠️  No hay más correos en espera de respuesta. Proceso detenido.")
                    break
                if i < 2:
                    print("\n--- Esperando 5 segundos para el siguiente lote... ---")
                    time.sleep(5)
            print("\n✅ Proceso de verificación de respuestas completado.")
        elif opcion == "5":
            print("¡Hasta pronto!")
            break
        else:
            print("Opción no válida. Intenta de nuevo.")

        input("\nPresiona Enter para continuar...")

if __name__ == "__main__":
    main()